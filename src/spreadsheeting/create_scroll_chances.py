import datetime

from openpyxl.styles import PatternFill

from . import open_workbook, open_sheet, save_workbook
from . import SpreadsheetPage

from ..config.config import date_format


from ..data.scrolls import Scroll, SCROLL_PROBABILITIES
from ..data.caskets import Casket
from ..data.casket_storage import load_values as load_casket_data
from ..data.scroll_chance_storage import load_values as load_scroll_chances_data
from ..data.jar_storage import load_values as load_jar_data

from ..calculations.calculation_strategies import CALCULATION_STRATEGIES, CalculationStrategy

from ..helpers.helpers import format_price

def create_scroll_spreadsheet():
    highest_margin = (0, 0)

    wb = open_workbook()
    scroll_sheet = open_sheet(wb, SpreadsheetPage.SCROLLS)

    casket_data = load_casket_data()
    scroll_chances = load_scroll_chances_data()
    jar_data = load_jar_data()

    scroll_sheet.cell(column=1, row=3, value="Spreadsheet Calculated:")
    scroll_sheet.cell(column=1, row=4, value=f'"{datetime.datetime.now().strftime(date_format)}"')

    red_fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
    green_fill = PatternFill(fill_type='solid', start_color='00FF00', end_color='00FF00')
    light_blue_fill = PatternFill(fill_type='solid', start_color='99CCFF', end_color='99CCFF')
    colors = {
        CalculationStrategy.Naive: light_blue_fill,
        CalculationStrategy.NBD: green_fill
    }

    row = 6
    row_sections = len(list(Scroll))

    best_profits = {}

    for scroll_index, scroll in enumerate(Scroll):
        scroll_sheet.cell(column=1, row=row, value=scroll.name + " SCROLL")
        row += 1

        if scroll not in best_profits.keys():
            best_profits[scroll] = {}

        scroll_sheet.cell(column=2, row=row, value="Number of Master Scrolls Wanted")
        row += 1

        master_scrolls_desired = 100
        current_scrolls_desired_value = scroll_sheet.cell(column=2, row=row)
        if current_scrolls_desired_value.value is None:
            scroll_sheet.cell(column=2, row=row, value=master_scrolls_desired)
        else:
            master_scrolls_desired = current_scrolls_desired_value.value

        row += 2

        possible_implings = []
        for impling, scroll_type in SCROLL_PROBABILITIES.items():
            if scroll in scroll_type.keys():
                possible_implings.append(impling)

        scroll_sheet.cell(column=2, row=row, value="Impling")
        scroll_sheet.cell(column=3, row=row, value="Price")
        scroll_sheet.cell(column=4, row=row, value="Loot Reward")
        scroll_sheet.cell(column=5, row=row, value="Casket Reward")
        scroll_sheet.cell(column=6, row=row, value="Scroll Probability")

        column = 6
        for column_index, calculation_strategy in enumerate(CALCULATION_STRATEGIES.keys()):
            scroll_sheet.cell(column=column + (column_index * 5), row=row, value=f"Expected Jar Count ({calculation_strategy.name})")
            scroll_sheet.cell(column=column + (column_index * 5) + 1, row=row, value=f"Expected Cost ({calculation_strategy.name})")
            scroll_sheet.cell(column=column + (column_index * 5) + 2, row=row, value=f"Expected Gross ({calculation_strategy.name})")
            scroll_sheet.cell(column=column + (column_index * 5) + 3, row=row, value=f"Expected Margin ({calculation_strategy.name})")
            scroll_sheet.cell(column=column + (column_index * 5) + 4, row=row, value=f"Margin % ({calculation_strategy.name})")

        format_row = row + 1
        for impling in possible_implings:
            row += 1

            print(f"Calculating {scroll} & {impling}")

            price = format_price(jar_data[impling.name]["price"])
            loot_reward = format_price(jar_data[impling.name]["loot"])

            casket_reward = format_price(casket_data[[casket.name for casket in Casket if scroll.name == casket.name][0]]["reward"])

            scroll_sheet.cell(column=2, row=row, value=impling.name)
            scroll_sheet.cell(column=3, row=row, value=price)
            scroll_sheet.cell(column=4, row=row, value=loot_reward)
            scroll_sheet.cell(column=5, row=row, value=casket_reward)
            scroll_sheet.cell(column=6, row=row, value=SCROLL_PROBABILITIES[impling][scroll])

            for column_index, calculation_strategy in enumerate(CALCULATION_STRATEGIES.keys()):

                if calculation_strategy not in best_profits[scroll].keys():
                    best_profits[scroll][calculation_strategy] = (0,0)

                base_scroll_chance = [x for x in scroll_chances[impling.name] if scroll.name.lower() in x[0]]
                base_scroll_chance = [y for x in base_scroll_chance for y in x[1].replace('"', '').split("/")]
                base_scroll_chance = int(base_scroll_chance[0]) / int(base_scroll_chance[1])

                master_scroll_chance = SCROLL_PROBABILITIES[impling][scroll]

                calculation_function = CALCULATION_STRATEGIES[calculation_strategy]

                scrolls_produced, expected_jars_needed, expected_cost = (
                    calculation_function(master_scrolls_desired, master_scroll_chance, base_scroll_chance, price)
                )

                gross_profit = \
                    expected_jars_needed * loot_reward + \
                    scrolls_produced * format_price(casket_data[scroll.name]["reward"]) +\
                    master_scrolls_desired * format_price(casket_data[Casket.MASTER.name]["reward"])

                margin = gross_profit - expected_cost

                scroll_sheet.cell(column=column + (column_index * 5), row=row, value=expected_jars_needed)
                scroll_sheet.cell(column=column + (column_index * 5) + 1, row=row, value=expected_cost)
                scroll_sheet.cell(column=column + (column_index * 5) + 2, row=row, value=gross_profit)
                margin_cell = scroll_sheet.cell(column=column + (column_index * 5) + 3, row=row, value=f"{margin:.2f}")
                margin_percent_cell = scroll_sheet.cell(column=column + (column_index * 5) + 4, row=row, value=f"{margin/gross_profit * 100:.1f}%")

                margin_cell.fill = red_fill
                margin_percent_cell.fill = red_fill
                if margin > 0:
                    margin_cell.fill = colors[calculation_strategy]
                    margin_percent_cell.fill = colors[calculation_strategy]

                if margin > best_profits[scroll][calculation_strategy][1]:
                    best_profits[scroll][calculation_strategy] = (impling.name, margin)

        for target_row in range(format_row, row+1):
            target_cell = scroll_sheet.cell(column=2, row=target_row)
            target_cell.fill = red_fill
            for cell_scroll, strategies in best_profits.items():
                if scroll != cell_scroll:
                    continue

                for strategy, value in strategies.items():
                    if value[1] == 0:
                        continue
                    if value[0] != target_cell.value:
                        continue
                    if value[1] > highest_margin[1]:
                        highest_margin = value
                    target_cell.fill = colors[strategy]
                    target_cell.value = f"{target_cell.value} (Best {strategy.name})"

        row += 3

    scroll_sheet.cell(column=2, row=3, value=f"HIGHEST MARGIN:")
    scroll_sheet.cell(column=3, row=3, value=highest_margin[0])
    scroll_sheet.cell(column=4, row=3, value=highest_margin[1])


    save_workbook(wb)
