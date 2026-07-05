import os
from enum import Enum, auto

from openpyxl import load_workbook, Workbook

SPREADSHEET_FILE = "clue_scroll.xlsx"

class SpreadsheetPage(Enum):
    IMPLINGS = auto()
    LOOT_ITEMS = auto()
    SCROLLS = auto()


def open_workbook():
    if os.path.exists(SPREADSHEET_FILE):
        return load_workbook(SPREADSHEET_FILE)

    return Workbook()


def save_workbook(wb):
    wb.save(SPREADSHEET_FILE)


def open_sheet(wb, page):
    if page not in SpreadsheetPage:
        raise ValueError('Page must be one of {}'.format(SpreadsheetPage))
    try:
        return wb[page.name]
    except KeyError:
        return wb.create_sheet(page.name, page.value-1)